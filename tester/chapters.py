import sys
import openpyxl as excel

def get_chapters():
    # SEARCH LOOP
    # scans the excel file to find the chapters of the given url
    filename = 'example_manga_list.xlsx'
    wb = excel.load_workbook(filename)
    sheet = wb.active
    # the name of the manga we want to download
    manga_url = sys.argv[1]
    manga_chapters = -1
    # search loop
    for i in range(2, 4588):
        match = sheet.cell(row=i, column=2).value

        #checks if the manga was found and breaks the loop
        if match == manga_url :
            print('---Manga Found---')
            manga_name = sheet.cell(row=i, column=1).value
            manga_chapters = sheet.cell(row=i, column=3).value
            print('Name:',manga_name)
            print('Chapters: ', str(manga_chapters))
            # if finds the given url breaks the loop
            break
    
    if manga_chapters == -1:
        print('Manga Not Found')
        exit()

    return manga_chapters


manga_chapters = get_chapters()
