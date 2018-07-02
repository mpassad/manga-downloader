from bs4 import BeautifulSoup
import requests
import sys
import os
import time
import ctypes
arg_len = len(sys.argv)
exec_file = sys.argv[0]
manga_url = sys.argv[1]

# clear screen console
def clear_console():
    os.system('cls' if os.name == 'nt' else 'reset')
# prints the manual of the app
def helper():
    print('')
    print('Help')
    print('--------------------------------------------------------------------------------------------')
    print('1) Download all the chapters of a manga pass only the url of the manga')
    print('# python downloader.py <manga-url>')
    print('--------------------------------------------------------------------------------------------')
    print('2) Download from chapter-1 to a certain chapter do as the following example')
    print('# python downloader.py <manga-url> -t <chapter-number>')
    print('--------------------------------------------------------------------------------------------')
    print('3) Download from a certain chapter till the end of the manga ')
    print('# python downloader.py <manga-url> -f <chapter-number>')
    print('--------------------------------------------------------------------------------------------')
    print('4) Download from a certain chapter till the end of the manga just')
    print('# python downloader.py <manga-url> -f <from_chapter> -t <to_chapter>')
    print('')
    time.sleep(10)
    exit()


# python downloader.py <manga_url> -f <from_chapter> -t <to_chapter>
#               0                       1             2         3               4       5

ctypes.windll.kernel32.SetConsoleTitleW("Manga Downloader V1.0")

# -c
#will download a certain chapter of the comic
def load_manga_list(from_chapter, to_chapter):
    pass


def load_chapters():
    pass


# gets as parameters the chapters from and to 
def make_dirs(from_chapter=1, to_chapter):
    # /some-path/manga_list/one-piece/chapter-1
    WORKING_DIR = os.path.dirname(os.path.realpath(__file__))
    LIBRARY_DIR= os.path.join(WORKING_DIR, 'LIBRARY') # /manga/one-piece
    MANGA_DIR = os.path.join(LIBRARY_DIR, sys.argv[1])
    # checks if a directory exists
    if not os.path.exists(LIBRARY_DIR):
        # if not exists creates it
        os.makedirs(LIBRARY_DIR)
    
    if not os.path.exists(MANGA_DIR):
        os.makedirs(MANGA_DIR)
    


def img_to_pdf():
    pass


clear_console()
print('MANGA DONWLOADER V1.0')  # message to inform the user which app use


# checks for the len of the arguments
if arg_len == 1:
    helper()
elif arg_len == 2:
    print('----Manga----')
    print('-URL: ', manga_url)
    print('-Name:')
    print('-Number of Chapters: ')
    
elif arg_len == 3:
    chapter = sys.argv[-1]
    print('Manga_url',)

# THE WORKING DIRECTORY


# import excel library
import openpyxl as excel
print('Working Directory: ', WORKING_DIR)
wb = excel.load_workbook('example_manga_list.xlsx')
sheets = wb.sheetnames
# holds the active sheet
sheet = wb.active
# search for a manga in the xlsx
url = sys.argv[1]
for i in range(1, 4588):
    match = sheet.cell(row=i, column=2).value
    if url == match:
        manga_chapters = sheet.cell(row=i, column=3).value
        print('Chapters: ', str(manga_chapters))
        print(type(manga_chapters))
        break

#creates a folder for every chapter for the manga we want to download
for i in range(1, manga_chapters + 1):
    chapter_num = 'CHAPTER-'+str(i)
    CHAPTER_PATH = os.path.join(SAVE_PATH, chapter_num)
    if not os.path.exists(CHAPTER_PATH):
        os.makedirs(CHAPTER_PATH)


# LOAD LOOP AND DOWNLOAD LOOP
# chapters loop
for j in range(1, manga_chapters+1):
    chapter_num = 'CHAPTER-'+str(j)
    CHAPTER_PATH = os.path.join(SAVE_PATH, chapter_num)
    if not os.path.exists(CHAPTER_PATH):
        os.makedirs(CHAPTER_PATH)
    # pages loop
    link = 'https://www.mangareader.net' + '/' + sys.argv[1] + '/' + str(j)
    r = requests.get(link)  # makes a request to the link
    data = r.text  # converts the request to text
    # makes a beautiful soup object and retrieves data from the link
    soup = BeautifulSoup(data, 'html.parser')
    img_tag = soup.find('img', {'id': 'img'})
    select_tag = soup.find('select', {'id': 'pageMenu'})
    

    # checks if the select tag doesnt exists
    if select_tag != None:
        children_tag = select_tag.find_all('option')
        num_of_pages = int(children_tag[-1].contents[0])
        txt = sys.argv[1] + '| Chapter ' + str(j) 
        print(txt)
        print(num_of_pages) 

        for i in range(1, num_of_pages+1):
            print('CHAPTER:',str(j))
            print('PAGE:',str(i))
            print('---------------')
            link = 'https://www.mangareader.net' +'/' +sys.argv[1] + '/' + str(j) + '/' + str(i)
            print(link)
            r = requests.get(link)  # makes a request to the link
            data = r.text  # converts the request to text
            soup = BeautifulSoup(data, 'html.parser')
            img_tag = soup.find('img', {'id': 'img'})
            source = img_tag['src']
            filename = str(j) + '-' + str(i) + '.jpg'
            # writes the image from link to file and saves it
            f = open(CHAPTER_PATH+'/'+filename, 'wb')
            f.write(requests.get(source).content)
            f.close()
    

    
