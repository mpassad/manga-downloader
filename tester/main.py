from bs4 import BeautifulSoup
import requests
import time
import sys

starting_time = time.time()
# this is the base link of all the manga
link = 'https://www.mangareader.net/alphabetical'
r = requests.get(link) # makes a request to the link
data = r.text # converts the request to text

# makes a beautiful soup object and retrieves data from the link
soup = BeautifulSoup(data, 'html.parser')
ul = soup.find_all('ul','series_alpha')

# creates an object with all the categories
main_list = soup.find_all('ul','series_alpha')
print('All the lists')
print('----------------')
main_list_children = main_list[0].findChildren()
all_manga_titles = []
urls = [] # empty list

# load data from the webpage
def load_data():
    for item in ul:  # loop to search to all the ul lists
        grandchilds = item.find_all('a')  # gets
        for item in grandchilds:
            #### print(item)
            urls.append(item['href'])  # append item to list urls
            # prints only the url for this manga not the base dir
            print(item['href'])
            cleared = item.contents[0]  # returns only the string from the tag
            all_manga_titles.append(cleared)  # stores the name to the list
            #print(cleared)
        print('======================================')

#save to xlsx
def save_xls():
    #loads the library to manipulate the xls file
    import openpyxl as excel
    workbook = excel.Workbook()
    sheet = workbook.active
    sheet.title = 'manga List'
    # this is the header of the table
    sheet.cell(row=1, column=1).value = 'Name'  # first column name
    sheet.cell(row=1, column=2).value = 'Url'  # second column name
    sheet.cell(row=1, column=3).value = 'Chapters'  # third column name

    # stores the manga titles in the workbook
    i = 2
    for item in all_manga_titles:
        sheet.cell(row=i, column=1).value = item
        i = i+1

    # stores the urls in workbook
    i = 2
    for item in urls:
        sheet.cell(row=i, column=2).value = item
        i = i+1

    #save the workbook with the following name
    workbook.save('example_manga_list.xlsx')

# function to load manga from the given url
def load_manga():
    BASE_DIR = 'https://www.mangareader.net'
    MANGA_DIR = '/boruto'
    CHAPTER_DIR = '/1'
    PAGE_DIR = '/1'
    # joins all the strings to a new string with the name FULL_DIR
    FULL_DIR = BASE_DIR + MANGA_DIR + CHAPTER_DIR + PAGE_DIR
    print(FULL_DIR)

# loads the TABLE OF CONTENTS for every manga
def manga_tbc(MANGA_DIR):
    BASE_DIR = 'https://www.mangareader.net'
    FULL_DIR = BASE_DIR+MANGA_DIR
    print('loading: ',FULL_DIR)
    r = requests.get(FULL_DIR)
    data = r.text # converts the object to text to load from beautiful soup
    soup = BeautifulSoup(data, 'html.parser')
    manga_table = soup.find('table',{'id': 'listing'})
    print(manga_table)

load_data()
save_xls()

print('Counters-----------')
print('count all mangas:', len(all_manga_titles))
print('count all urls:', len(urls))

#END_OF_FILE


#load the xls
#read all the links from the file
# make a soup object
# load the link of every manga
# read the chapters in each url
# store the data to list
# write the list to workbook 
#save the workbook
#close it

def count_chapters():
    chapters = []
    sum = 0
    for item in urls:
        print(item)
        link = 'https://www.mangareader.net'+item
        r = requests.get(link)  # makes a request to the link
        data = r.text  # converts the request to text
        bt = sys.getsizeof(data)
        #print(str(bt), 'bytes')
        # makes a beautiful soup object and retrieves data from the link
        soup = BeautifulSoup(data, 'html.parser')
        
        table = soup.find('table', {'id': 'listing'})
        children = table.findChildren()
        sum = sum + bt
        #print(sum)
        if len(children)>4:
            a_tag = table.find_all('a')
            del table  # deletes and free up space in memory
            text_chapters = a_tag[-1].contents[0]
            del a_tag
            breaker = text_chapters.split(' ')
            del text_chapters
            num_chapters = int(breaker[-1])
            chapters.append(num_chapters)
            del breaker
            del soup
        else:
            #print('chapters 0')
            chapters.append(0)
        
        #print(0)
    print('Sum of bytes')
    print(sum)
    return chapters

num_chapters = count_chapters()

import openpyxl as excel
wb = excel.load_workbook('example_manga_list.xlsx')
sheet = wb.active

i=2
for item in num_chapters:
    sheet.cell(row=i , column=3).value = item
    i = i+1

wb.save('example_manga_list.xlsx')
print('ENDED')
ending_time = time.time()
print('Time lenght:', str(ending_time-starting_time), 'seconds') # bench marking

