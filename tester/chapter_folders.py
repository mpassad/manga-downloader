import os
import sys

# THE WORKING DIRECTORY
BASE_PATH = os.path.dirname(os.path.realpath(__file__))  
# /some-path/manga_list/one-piece/chapter-1
# /some-path/manga_list/one-piece/chapter-2
# /some-path/manga_list/one-piece/chapter-3
MANGA_DIR = 'manga_list'
MANGA_LIST_PATH = os.path.join(BASE_PATH, MANGA_DIR)
# checks if a directory exists
if not os.path.exists(MANGA_LIST_PATH):
    os.makedirs(MANGA_LIST_PATH)
# manga base directory that contains all the paths
SAVE_PATH = os.path.join(MANGA_LIST_PATH, sys.argv[1])

if not os.path.exists(SAVE_PATH):
    os.makedirs(SAVE_PATH)

# import excel library
import openpyxl as excel
print('Working Directory: ', BASE_PATH)
wb = excel.load_workbook('example_manga_list.xlsx')
sheets = wb.sheetnames
# holds the active sheet
sheet = wb.active
# search for a manga in the xlsx
url = sys.argv[1]
for i in range(1, 4588):
    match = sheet.cell(row=i, column=2).value
    if url == match:
        manga_chapters = sheet.cell(row=i, column=3).value
        print('Chapters: ', str(manga_chapters))
        print(type(manga_chapters))
        break

for i in range(1,manga_chapters + 1):
    chapter_num = 'CHAPTER-'+str(i)
    CHAPTER_PATH = os.path.join(SAVE_PATH, chapter_num)
    if not os.path.exists(CHAPTER_PATH):
        os.makedirs(CHAPTER_PATH)

    
